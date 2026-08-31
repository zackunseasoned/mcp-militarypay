"""Parser for the DTMO "All BAH Rates" Excel workbook.

This is the other bulk download offered beside the ASCII bundle, and it is the
only published form the off-cycle adjustments appear in - there is no separate
ASCII bundle for, say, the 2026 TX270 temporary increase.

The workbook is a clean grid, not the government-Excel horror the build spec
braced for: two sheets, `With` and `Without`, a title in row 1, a header in row
2, and one row per Military Housing Area after that.

    MHA | MHA_NAME | E01..E09 | W01..W05 | O01E O02E O03E | O01..O07

Two things differ from the ASCII rate files:

* There is **no ZIP-to-MHA crosswalk** here. A rate set ingested from the
  workbook therefore cannot be an annual baseline on its own; ZIP resolution
  goes through whichever annual set does carry the crosswalk.
* The workbook publishes **24** rate columns, ending at O07, where the ASCII
  files carry 27. The three extra ASCII columns repeat the O-7 value - DTMO
  treats O-7 and above as one "O-7/O-7+" bucket. Rows read from here are
  expanded the same way so both sources yield the same grade set.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

from ..sources import BAH_SENIOR_OFFICER_GRADES

class ParseError(ValueError):
    """The workbook did not have the expected shape."""


WITH_DEPENDENTS_SHEET = "With"
WITHOUT_DEPENDENTS_SHEET = "Without"

_MHA_RE = re.compile(r"^[A-Z]{2}\d{3}$")

# Workbook column label -> canonical pay grade.
_GRADE_LABEL_RE = re.compile(r"^([EWO])0?(\d{1,2})(E)?$", re.IGNORECASE)


def normalize_grade_label(label: str) -> str | None:
    """'E01' -> 'E-1', 'O01E' -> 'O-1E', 'W05' -> 'W-5'."""
    match = _GRADE_LABEL_RE.match((label or "").strip())
    if not match:
        return None
    branch, number, prior_enlisted = match.groups()
    grade = f"{branch.upper()}-{int(number)}"
    return grade + "E" if prior_enlisted else grade


@dataclass
class BahWorkbookSheet:
    with_dependents: bool
    rates: dict[tuple[str, str], float] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    @property
    def mha_codes(self) -> set[str]:
        return {mha for mha, _ in self.rates}


@dataclass
class BahWorkbook:
    mha_names: dict[str, str] = field(default_factory=dict)
    with_dependents: BahWorkbookSheet | None = None
    without_dependents: BahWorkbookSheet | None = None
    title: str | None = None
    warnings: list[str] = field(default_factory=list)


def _load_sheet(workbook, sheet_name: str, with_dependents: bool
                ) -> tuple[BahWorkbookSheet, dict[str, str]]:
    if sheet_name not in workbook.sheetnames:
        raise ParseError(
            f"no {sheet_name!r} sheet in the workbook; it has "
            f"{workbook.sheetnames}"
        )
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(min_row=1, values_only=True))
    if len(rows) < 3:
        raise ParseError(f"the {sheet_name!r} sheet has no data rows")

    # Row 1 is a title; the header is the first row whose first cell is 'MHA'.
    header_index = None
    for index, row in enumerate(rows[:5]):
        if row and str(row[0] or "").strip().upper() == "MHA":
            header_index = index
            break
    if header_index is None:
        raise ParseError(
            f"no header row starting with 'MHA' in the {sheet_name!r} sheet - "
            f"the workbook layout has probably changed"
        )

    header = rows[header_index]
    columns: dict[int, str] = {}
    for position, label in enumerate(header):
        grade = normalize_grade_label(str(label or ""))
        if grade:
            columns[position] = grade
    if not columns:
        raise ParseError(
            f"no pay grade columns recognised in the {sheet_name!r} header: "
            f"{[h for h in header if h]}"
        )

    sheet = BahWorkbookSheet(with_dependents=with_dependents)
    names: dict[str, str] = {}

    for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        if not row or not row[0]:
            continue
        mha = str(row[0]).strip().upper()
        if not _MHA_RE.match(mha):
            raise ParseError(
                f"{sheet_name!r} row {row_number}: {mha!r} is not an MHA code"
            )
        if len(row) > 1 and row[1]:
            names[mha] = " ".join(str(row[1]).split())

        for position, grade in columns.items():
            if position >= len(row):
                continue
            value = row[position]
            if value is None or str(value).strip() == "":
                sheet.warnings.append(
                    f"{sheet_name!r} row {row_number}: empty rate for {mha} {grade}"
                )
                continue
            try:
                sheet.rates[(mha, grade)] = float(
                    str(value).replace(",", "").replace("$", "").strip()
                )
            except ValueError as exc:
                raise ParseError(
                    f"{sheet_name!r} row {row_number}: rate for {mha} {grade} "
                    f"is not a number: {value!r}"
                ) from exc

    if not sheet.rates:
        raise ParseError(f"the {sheet_name!r} sheet contained no rate rows")

    # DTMO collapses O-7 and above into one bucket. The workbook stops at O-7
    # while the ASCII files repeat its value out to O-10; expand here so a
    # lookup for O-8 against a workbook-sourced rate set does not miss and fall
    # through to a stale annual rate.
    expanded = 0
    for mha in sheet.mha_codes:
        anchor = sheet.rates.get((mha, "O-7"))
        if anchor is None:
            continue
        for grade in BAH_SENIOR_OFFICER_GRADES:
            if (mha, grade) not in sheet.rates:
                sheet.rates[(mha, grade)] = anchor
                expanded += 1
    if expanded:
        sheet.warnings.append(
            f"expanded the O-7 rate to {sorted(set(BAH_SENIOR_OFFICER_GRADES) - {'O-7'})} "
            f"for {expanded // 3} MHAs (DTMO's 'O-7/O-7+' bucket)"
        )
    return sheet, names


def parse_bah_workbook(xlsx_bytes: bytes) -> BahWorkbook:
    """Parse a DTMO BAH rates workbook into both dependency statuses."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ParseError(
            "reading the BAH Excel workbook needs openpyxl: pip install openpyxl"
        ) from exc

    workbook = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), read_only=True, data_only=True
    )
    try:
        result = BahWorkbook()
        first = workbook[workbook.sheetnames[0]]
        for row in first.iter_rows(min_row=1, max_row=1, values_only=True):
            if row and row[0]:
                result.title = " ".join(str(row[0]).split())
            break

        with_sheet, with_names = _load_sheet(
            workbook, WITH_DEPENDENTS_SHEET, with_dependents=True
        )
        without_sheet, without_names = _load_sheet(
            workbook, WITHOUT_DEPENDENTS_SHEET, with_dependents=False
        )
    finally:
        workbook.close()

    result.with_dependents = with_sheet
    result.without_dependents = without_sheet
    result.mha_names = {**without_names, **with_names}

    if with_sheet.mha_codes != without_sheet.mha_codes:
        only_with = sorted(with_sheet.mha_codes - without_sheet.mha_codes)[:5]
        only_without = sorted(without_sheet.mha_codes - with_sheet.mha_codes)[:5]
        result.warnings.append(
            f"the two sheets cover different MHAs: only-with {only_with}, "
            f"only-without {only_without}"
        )
    result.warnings.extend(with_sheet.warnings[:3])
    result.warnings.extend(without_sheet.warnings[:3])
    return result


def diff_workbooks(baseline: BahWorkbook, updated: BahWorkbook
                   ) -> dict[str, set[str]]:
    """MHAs whose rates differ between two workbooks, by dependency status.

    An off-cycle publication is the annual workbook with a handful of areas
    changed, so this is how the affected set is identified rather than trusting
    a filename.
    """
    changes: dict[str, set[str]] = {"with_dependents": set(), "without_dependents": set()}
    for key, attr in (("with_dependents", "with_dependents"),
                      ("without_dependents", "without_dependents")):
        old = getattr(baseline, attr)
        new = getattr(updated, attr)
        if old is None or new is None:
            continue
        for (mha, grade), rate in new.rates.items():
            if old.rates.get((mha, grade)) != rate:
                changes[key].add(mha)
    return changes
