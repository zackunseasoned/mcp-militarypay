"""Tests for the DTMO BAH rates workbook.

This is the other bulk download, and the only published form the off-cycle
adjustments appear in — there is no separate ASCII bundle for the 2026 TX270
temporary increase.
"""

import pytest

from mcp_militarypay.parsers.bah_xlsx import (
    ParseError,
    diff_workbooks,
    normalize_grade_label,
    parse_bah_workbook,
)


@pytest.mark.parametrize(
    "label,expected",
    [("E01", "E-1"), ("E09", "E-9"), ("W05", "W-5"),
     ("O01E", "O-1E"), ("O03E", "O-3E"), ("O01", "O-1"), ("O07", "O-7"),
     ("MHA", None), ("MHA_NAME", None), ("", None)],
)
def test_normalize_grade_label(label, expected):
    assert normalize_grade_label(label) == expected


class TestWorkbook:
    def test_reads_both_sheets_and_names(self, bah_workbook_bytes):
        book = parse_bah_workbook(bah_workbook_bytes)
        assert book.title.startswith("2026 BAH Rates")
        assert book.mha_names["TX270"] == "ABILENE/DYESS AFB, TX"
        assert book.with_dependents.mha_codes == book.without_dependents.mha_codes

    def test_with_dependents_exceeds_without(self, bah_workbook_bytes):
        book = parse_bah_workbook(bah_workbook_bytes)
        for key, rate in book.with_dependents.rates.items():
            assert rate > book.without_dependents.rates[key]

    def test_expands_the_o7_bucket_to_o10(self, bah_workbook_bytes):
        """The workbook stops at O07 while the ASCII files repeat that value out
        to O-10. Without the expansion, an O-8 lookup against a workbook-sourced
        off-cycle set would miss and fall through to a stale annual rate."""
        book = parse_bah_workbook(bah_workbook_bytes)
        rates = book.with_dependents.rates
        for mha in book.with_dependents.mha_codes:
            anchor = rates[(mha, "O-7")]
            for grade in ("O-8", "O-9", "O-10"):
                assert rates[(mha, grade)] == anchor
        assert any("O-7/O-7+" in w for w in book.warnings)

    def test_covers_the_same_grade_set_as_the_ascii_files(self, bah_workbook_bytes):
        from mcp_militarypay.sources import BAH_RATE_COLUMNS

        book = parse_bah_workbook(bah_workbook_bytes)
        grades = {grade for _, grade in book.with_dependents.rates}
        assert grades == set(BAH_RATE_COLUMNS)

    def test_a_missing_sheet_is_reported(self, tmp_path):
        import openpyxl

        workbook = openpyxl.Workbook()
        workbook.active.title = "With"
        workbook.active.cell(1, 1, "title")
        workbook.active.cell(2, 1, "MHA")
        workbook.active.cell(2, 3, "E01")
        workbook.active.cell(3, 1, "TX270")
        workbook.active.cell(3, 3, 100)
        path = tmp_path / "one_sheet.xlsx"
        workbook.save(path)
        with pytest.raises(ParseError, match="Without"):
            parse_bah_workbook(path.read_bytes())

    def test_a_missing_header_row_is_reported(self, tmp_path):
        import openpyxl

        workbook = openpyxl.Workbook()
        for name in ("With", "Without"):
            sheet = workbook.create_sheet(name)
            sheet.cell(1, 1, "some title")
            sheet.cell(2, 1, "NOT_THE_HEADER")
            sheet.cell(3, 1, "TX270")   # enough rows to reach the header check
            sheet.cell(4, 1, "CA606")
        del workbook["Sheet"]
        path = tmp_path / "no_header.xlsx"
        workbook.save(path)
        with pytest.raises(ParseError, match="MHA"):
            parse_bah_workbook(path.read_bytes())


class TestDiff:
    def test_identifies_only_the_changed_area(
        self, bah_workbook_bytes, bah_workbook_increase_bytes
    ):
        """An off-cycle publication is the full annual table with a few areas
        changed, so the affected set is derived rather than taken on trust."""
        baseline = parse_bah_workbook(bah_workbook_bytes)
        updated = parse_bah_workbook(bah_workbook_increase_bytes)
        changed = diff_workbooks(baseline, updated)
        assert changed["with_dependents"] == {"TX270"}
        assert changed["without_dependents"] == {"TX270"}

    def test_identical_workbooks_show_no_change(self, bah_workbook_bytes):
        book = parse_bah_workbook(bah_workbook_bytes)
        changed = diff_workbooks(book, book)
        assert changed["with_dependents"] == set()
